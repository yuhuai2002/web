"""
炜易达海运清关资料
"""
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from unzip import get_zip_filename, unzip_file, delete_all_files
import glob
import re
import os
import shutil
import pandas as pd
import pickle
# from sqlalchemy import create_engine, Table, MetaData
import traceback
import datetime
from urllib.parse import quote_plus


# --------------------------
# 常量定义
"""
work_dir: zip压缩包接收到后保存的位置
extracted_path: 解压缩后文件保存的位置
output_dir: 处理后CSV文件保存的位置,在CSV文件生成后自动将该文件提供浏览器下载到本地
source: CSV模板文件所在路径
"""
# work_dir = r"D:\垃圾堆\文档\物流\parcel\zip"
# extracted_path = r"D:\垃圾堆\文档\物流\parcel\unzip"
# output_dir = r"D:\垃圾堆\文档\物流\parcel\output"
# source = r"D:\垃圾堆\文档\物流\parcel\tmplate.xlsx" 
work_dir = "/app/www/zip" 
extracted_path = "/app/www/unzip" 
output_dir = "/app/www/output" 
source = "/app/xgo/tmplate.xlsx" 
# delete_all_files(work_dir)

it_list = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 
           'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2', 'P2']
headers = ["Mawb", "Order No", "CTN MARKS", "Reference", "Reference 2", "Internal Account Number",
           "Shipper Name", "Shipper Address 1", "Shipper Address 2", "Shipper Address 3",
           "Shipper City", "Shipper State", "Shipper Postcode", "Shipper Country Code",
           "Shipper Turn", "Consignee Name", "Consignee Address1", "Consignee Address2",
           "Consignee Address3", "Consignee City", "Consignee State", "Consignee Postcode",
           "Consignee Country Code", "Consignee Email", "Consignee Telephone Number",
           "Consignee Turn", "Consignee VAT Reg No", "Deferment", "GROSS WEIGHT", "Pieces",
           "Weight UOM", "Item Value", "Net Weight", "Currency", "Incoterms",
           "Item Description", "Item HS Code", "Item Quantity", "Total Value",
           "Country Of Origin", "Vendor"]


underline_style = Side(style='thin', color='000000')

di = {}
rows = []
# ----------------------------------
# 函数定义

def get_info(ws,current_order_name):
    if ws["A1"].value.rstrip() == "clearance instruction":
        print("clearance instruction")
        mawb = ''                     # 柜号
        from_by = ws["A3"].value                    # 
        shipper_name = ws["E4"].value               # 发货者名称
        shipper_address = ws["E5"].value            # 发货者地址
        shipper_city = ws["E6"].value if ws["E6"].value != None else '' ###
        consignee_vat = ws["E7"].value if ws["E7"].value != None else ''             # 收货人增值税号
        consignee_eori = ws["E8"].value if ws["E8"].value != None else ''            # 收货人EORI号
        consignee_name = ws["E9"].value if ws["E9"].value != None else ''             # 收货人名称
        if ws["E10"].value != None:
            consignee_address = ws["E10"].value.split('\n')       # 收货人街道
            # consignee_address_1 = consignee_address[:-2]
            consignee_address_1 = ' '.join(consignee_address)
        else:
            consignee_address = ""
            consignee_address_1 = ""
        consignee_city = ws["E11"].value if ws["E11"].value != None else ''
        consignee_postcode = ws["E12"].value if ws["E12"].value != None else ''
        ctn_marks = "" ###
        shipper_postcode = "518100"                 # 发货人邮政编码
        shipper_country_code = "CN"                 # 发货人国家代码
        consignee_country_code = "GB"               # 收货人国家代码
        deferment = "PVA"
    if ws["A1"].value.rstrip() == "Commercial Invoice":
        print("Commercial Invoice")
        mawb = ''                     # 柜号
        from_by = ws["A3"].value                    # 
        shipper_name = ws["E4"].value               # 发货者名称
        shipper_address = ws["E5"].value            # 发货者地址
        shipper_city = ws["E6"].value if ws["E6"].value != None else ''
        consignee_vat = ws["E7"].value if ws["E7"].value != None else ''             # 收货人增值税号
        consignee_eori = ws["E8"].value if ws["E8"].value != None else ''            # 收货人EORI号
        consignee_name = ws["E9"].value if ws["E9"].value != None else ''            # 收货人名称
        if ws["E10"].value != None:
            consignee_address = ws["E10"].value.split('\n')       # 收货人街道
            consignee_address_1 = consignee_address[:-3]
            consignee_address_1 = ' '.join(consignee_address_1)
            consignee_city = consignee_address[-3]  ###
            consignee_postcode = consignee_address[-2] ###
        else:
            consignee_address = ""
            consignee_address_1 = ""
            consignee_city = ""
            consignee_postcode = ""
        ctn_marks = "" ###
        shipper_postcode = "518100"                 # 发货人邮政编码
        shipper_country_code = "CN"                 # 发货人国家代码
        consignee_country_code = "GB"               # 收货人国家代码
        deferment = "PVA"
        
    return [mawb,current_order_name,ctn_marks,'','','',shipper_name,
            shipper_address,'','',shipper_city,'',shipper_postcode,shipper_country_code,
            '',consignee_name,consignee_address_1,'','',consignee_city,'',consignee_postcode,
            consignee_country_code,'','',consignee_vat,consignee_vat,deferment
    ]


def copy_and_rename_file(source_path, destination_dir, new_name):
    if not os.path.exists(source_path):
        error_msg = f"错误: 源文件 '{source_path}' 不存在"
        print(error_msg)
        return False, error_msg
    if not os.path.isfile(source_path):
        error_msg = f"错误: '{source_path}' 不是文件"
        print(error_msg)
        return False, error_msg
    try:
        os.makedirs(destination_dir, exist_ok=True)
    except Exception as e:
        error_msg = f"创建目录失败: {str(e)}"
        print(error_msg)
        return False, error_msg
    destination_path = os.path.join(destination_dir, new_name)
    if os.path.exists(destination_path):
        print(f"警告: 目标文件 '{destination_path}' 已存在，将被覆盖")
    try:
        shutil.copy2(source_path, destination_path)
        print(f"成功: 文件已从 '{source_path}' 复制到 '{destination_path}'")
        return True, destination_path
    except Exception as e:
        error_msg = f"复制文件失败: {str(e)}"
        print(error_msg)
        return False, error_msg


def number_partition(numbers):
    total = sum(numbers)
    if total == 0:
        return tuple([0.0] * len(numbers))
    weights = [round(num / total, 2) for num in numbers]
    sum_weights = sum(weights)
    if sum_weights != 1.0:
        difference = round(1.0 - (sum_weights - weights[-1]), 2)
        weights[-1] = difference
    return weights


def get_mawb(extracted_path):
    pattern2 = r"B2B\s*Delivery\s*Instructions-\s*(.*?)\.xlsx"
    pattern3 = r"Delivery\s*Instructions-\s*(.*?)\.xlsx"
    try:
        file_list = glob.glob(os.path.join(extracted_path, "B2B  Delivery Instructions*"))
        print("1_file_list[0]: ", file_list[0])
        Mawb = re.search(pattern2, file_list[0]).group(1)
    except:
        file_list = glob.glob(os.path.join(extracted_path, "Delivery Instructions*"))
        print("2_file_list[0]: ", file_list[0])
        Mawb = re.search(pattern3, file_list[0]).group(1)
    
    Delivery_Instructions = file_list[0]
    new_name = f"{Mawb}---B2B  CSV.xlsx"
    _, output_csv_dir = copy_and_rename_file(source, output_dir, new_name)
    print("Mawb: ",Mawb)
    print("New output CSV Path: ", output_csv_dir)
    return Delivery_Instructions, Mawb, output_csv_dir


def open_delivery_instructions(Delivery_Instructions):
    wb_B2B_Delivery_Instructions = load_workbook(filename=Delivery_Instructions, data_only=True)
    ws_B2B_Delivery_Instructions = wb_B2B_Delivery_Instructions.active
    wb_B2B_Delivery_Instructions.close()
    return ws_B2B_Delivery_Instructions


def locate_column(ws_B2B_Delivery_Instructions):
    # 获取CTN_MARK和Delivery_Method所在的列位置
    CTN_MARK = """CTN
Mark"""
    Delivery_Method = """Delivery
Method"""
    for i, it in enumerate(it_list):
        if ws_B2B_Delivery_Instructions[it].value == CTN_MARK:
            CTN_pos = i
        if ws_B2B_Delivery_Instructions[it].value == Delivery_Method:
            Delivery_Method_pos = i
    return CTN_pos, Delivery_Method_pos


def read_delivery_instructions(CTN_pos, Delivery_Method_pos):
    global di
    global row
    # 获取CTN和Delivery_Method对应的值
    for row in list(ws_B2B_Delivery_Instructions.iter_rows(values_only=True))[2:]:
        if row[0] != None:
            di[row[0]] = {"ctn": row[CTN_pos], "vendor": row[Delivery_Method_pos]}


def process(Mawb, output_csv_dir):
    # 对所有文件逐个处理
    global di
    global row
    global extracted_path
    pattern = r"Invoice & Packing List/(.*)\.xlsx"           # Linux系统
    # pattern = r".*\\Invoice & Packing List\\(.*)\.xlsx"    # Windows系统
    
    
    for i, file_path in enumerate(sorted(glob.glob(os.path.join(extracted_path, 'Invoice & Packing List/*.xlsx')))):
        print(f"Processing file {i+1}: {file_path}")
        match = re.search(pattern, file_path)
        current_order_name = match.group(1)
        wb = load_workbook(filename=file_path, data_only=True)
        # ws = wb.active
        ws = wb.worksheets[0]
        info_1 = get_info(ws, current_order_name)
        info_1[2] = di[current_order_name]["ctn"]
        info_1[0] = Mawb

        headers_name = ws["A1"].value.rstrip() # 表头名称
        item = []
        cache = []
        flag = []
        head = -1
        swh = False
        j = 0
        p = 2

        if headers_name == 'Commercial Invoice':
            for row in ws.iter_rows(values_only=True):
                if type(row[0]) == int and row[2] != None:
                    gross_weight = row[9]
                    pieces = row[11]
                    weight_uom = "kg"
                    item_value = row[7]
                    currency = row[12]
                    incoterms = "CIF"
                    item_description = row[3]
                    hs_code = row[2]
                    item_quantity = row[6]
                    total_value = row[8]
                    country_of_origin = "CN"
                    vendor = di[current_order_name]["vendor"]
                    cache.append(gross_weight)

                    if pieces == None:
                        if swh == False:
                            head = j - 1
                        if head != -1:
                            flag[head] = 1
                        flag.append(p)
                        p += 1
                        swh = True
                    else:
                        flag.append(0)
                        p = 2
                        swh = False
                    j += 1

                    info_2 = [gross_weight,pieces,weight_uom,item_value,"",
                            currency,incoterms,item_description,hs_code,item_quantity,
                            total_value,country_of_origin,vendor]
                    # item.append(info_1 + info_2)
                    item.append(info_1 + info_2)
        if headers_name == 'clearance instruction':
            for row in ws.iter_rows(values_only=True):
                if type(row[0]) == int and row[2] != None:
                    gross_weight = row[9]
                    pieces = row[10]
                    weight_uom = "kg"
                    item_value = row[6]
                    currency = row[12]
                    incoterms = "CIF"
                    item_description = row[3]
                    hs_code = row[2]
                    item_quantity = row[5]
                    total_value = row[7]
                    country_of_origin = "CN"
                    vendor = di[current_order_name]["vendor"]
                    cache.append(gross_weight)

                    if pieces == None:
                        if swh == False:
                            head = j - 1
                        if head != -1:
                            flag[head] = 1
                        flag.append(p)
                        p += 1
                        swh = True
                    else:
                        flag.append(0)
                        p = 2
                        swh = False
                    j += 1

                    info_2 = [gross_weight,pieces,weight_uom,item_value,"",
                            currency,incoterms,item_description,hs_code,item_quantity,
                            total_value,country_of_origin,vendor]
                    # item.append(info_1 + info_2)
                    item.append(info_1 + info_2)
            
        pre = -1
        arg = []
        dump = []
        pieces_list = []

        for i, it in enumerate(flag):
            if it > pre and it != 0:
                arg.append(i)
            if it < pre and it != 0:
                for k in arg:
                    dump.append(cache[k])
                res = number_partition(tuple(dump))
                pieces_list.append(res)
                arg = []
                dump = []
                arg.append(i)
            if it < pre and it == 0:
                for k in arg:
                    dump.append(cache[k])
                res = number_partition(tuple(dump))
                pieces_list.append(res)
                arg = []
                dump = []
            pre = it

        if len(arg) > 0:
            for k in arg:
                dump.append(cache[k])
            res = number_partition(tuple(dump))
            pieces_list.append(res)

        res = []
        for i in pieces_list:
            for j in i:
                res.append(j)

        for i, j in enumerate(flag):
            if j != 0:
                item[i][29] = res.pop(0)

        for i in item:
            i[32] = round((i[28] - i[29] * 2)/ i[37], 3)
        
        rows.extend(item)

        
        # 将当前文件数据附加到输出的CSV文件中
        wb_target = load_workbook(filename=output_csv_dir)
        ws_target = wb_target.active
        for index, i in enumerate(item):
            print(i)
            ws_target.append(i)
            
        wb.close()
        wb_target.save(filename=output_csv_dir)
        wb_target.close()


def mark_yellow(output_csv_dir):
    global rows
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    NUM = len(rows)
    wb_target = load_workbook(filename=output_csv_dir)
    ws_target = wb_target.active
    for i in range(NUM):
        ws_target[f"AB{i + 2}"].fill = yellow_fill
    wb_target.save(filename=output_csv_dir)
    wb_target.close()
    # 保存到文件
    with open("/app/www/array/list.pkl", "wb") as f:
        pickle.dump(rows, f)

try:
    delete_all_files(extracted_path)
    delete_all_files(output_dir)
    zip_file = get_zip_filename(work_dir)
    extract_files = unzip_file(zip_file, extracted_path)

    Delivery_Instructions, Mawb, output_csv_dir = get_mawb(extracted_path)
    ws_B2B_Delivery_Instructions = open_delivery_instructions(Delivery_Instructions)
    CTN_pos, Delivery_Method_pos = locate_column(ws_B2B_Delivery_Instructions)
    read_delivery_instructions(CTN_pos, Delivery_Method_pos)
    process(Mawb, output_csv_dir)
    mark_yellow(output_csv_dir)
    
    delete_all_files(work_dir)
    print("✅【成功】所有文件处理完成，输出文件已保存。")

except Exception as e:
    print("❌【错误】处理过程中发生异常，原因如下：")
    print(str(e))
    print("详细错误堆栈：")
    print(traceback.format_exc())
    delete_all_files(work_dir)



"""
rows = [
    {
        'Mawb': row[0],
        'Order_No': row[1],
        'CTN_MARKS': row[2],
        'Reference': row[3],
        'Reference_2': row[4],
        'Internal_Account_Number': row[5],
        'Shipper_Name': row[6],
        'Shipper_Address_1': row[7],
        'Shipper_Address_2': row[8],
        'Shipper_Address_3': row[9],
        'Shipper_City': row[10],
        'Shipper_State': row[11],
        'Shipper_Postcode': row[12],
        'Shipper_Country_Code': row[13],
        'Shipper_Turn': row[14],
        'Consignee_Name': row[15],
        'Consignee_Address1': row[16],
        'Consignee_Address2': row[17],
        'Consignee_Address3': row[18],
        'Consignee_City': row[19],
        'Consignee_State': row[20],
        'Consignee_Postcode': row[21],
        'Consignee_Country_Code': row[22],
        'Consignee_Email': row[23],
        'Consignee_Telephone_Number': row[24],
        'Consignee_Turn': row[25],
        'Consignee_VAT_Reg_No': row[26],
        'Deferment': row[27],
        'GROSS_WEIGHT': row[28],
        'Pieces': row[29],
        'Weight_UOM': row[30],
        'Item_Value': row[31],
        'Net_Weight': row[32],
        'Currency': row[33],
        'Incoterms': row[34],
        'Item_Description': row[35],
        'Item_HS_Code': row[36],
        'Item_Quantity': row[37],
        'Total_Value': row[38],
        'Country_Of_Origin': row[39],
        'Vendor': row[40]
    } for row in rows
]



# 确保日志目录存在
os.makedirs("/home/ubuntu/gyd/qianyi/flaskProject-trans/www", exist_ok=True)
log_file = "/home/ubuntu/gyd/qianyi/flaskProject-trans/www/import_log.txt"

def log_message(message):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_message = f"[{timestamp}] {message}\n"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(full_message)
    print(message)

try:
    log_message("【开始连接数据库】")
    log_message(f"Excel 路径: {mes}")

    db_config = {
        'username': 'root',
        'password': '#Sdf3@3&gh.dfYYs34Op7^3@#',
        'host': '59.110.156.56',
        'port': '3306',
        'database': 'fapiao'
    }

    encoded_password = quote_plus(db_config['password'])
    connection_string = (
        f"mysql+mysqlconnector://{db_config['username']}:{encoded_password}"
        f"@{db_config['host']}:{db_config['port']}/{db_config['database']}"
    )

    engine = create_engine(connection_string)
    metadata = MetaData()

# 反射获取表结构（假设表已存在）
    table_sc = Table('sc', metadata, autoload_with=engine)

    # 执行批量插入
    with engine.connect() as conn:
        conn.execute(table_sc.insert(), rows)
        conn.commit()

except Exception as e:
    log_message("❌【错误】写入数据库失败，原因如下：")
    log_message(str(e))
    log_message("详细错误堆栈：")
    log_message(traceback.format_exc())

delete_all_files(extracted_path)
delete_all_files(work_dir)

"""